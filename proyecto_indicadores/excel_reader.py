"""
excel_reader.py
─────────────────────────────────────────────────────────────────────────────
Funciones de lectura de archivos Excel:
  - Leer el maestro
  - Leer la fecha desde la celda configurada (o escaneo automático de la hoja)
  - Leer la tabla de datos con auto-detección del encabezado
  - Normalizar nombres de columnas para tolerancia de tildes y caracteres extra
  - Construir los campos Fecha, Anio, Mes y Periodo_YYYYMM
─────────────────────────────────────────────────────────────────────────────
"""

import logging
import re
from datetime import datetime, date
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter, column_index_from_string

from config import (
    DATE_CELL,
    MAESTRO_SHEET,
    INDIVIDUAL_SHEET,
    INDIVIDUAL_HEADER_ROW,
    MES_FORMAT,
)
from normalization import normalize_value

logger = logging.getLogger("indicadores")

# Mapa de número de mes → nombre en español (para MES_FORMAT = "texto")
MESES_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
}


# ─────────────────────────────────────────────────────────────
# MAESTRO
# ─────────────────────────────────────────────────────────────

def read_maestro(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Maestro no encontrado: {path}")

    logger.info(f"Leyendo maestro: {path.name}")
    sheet = MAESTRO_SHEET if MAESTRO_SHEET is not None else 0
    df = pd.read_excel(path, sheet_name=sheet, header=0, engine="openpyxl")

    if df.empty:
        raise ValueError(f"El maestro está vacío: {path.name}")

    logger.info(f"  → {len(df)} registros cargados del maestro.")
    return df


# ─────────────────────────────────────────────────────────────
# UTILIDADES DE FECHA
# ─────────────────────────────────────────────────────────────

def _parse_date_value(raw_value) -> datetime | None:
    """
    Convierte cualquier valor de celda en datetime.
    Retorna None si no es posible, sin lanzar excepción.

    Soporta adicionalmente:
      - "03/2026" o "3/2026"  → primer día del mes (MM/YYYY)
      - "2026-03"             → primer día del mes (YYYY-MM)
      - Cualquier texto que contenga esos patrones
    """
    if raw_value is None:
        return None

    if isinstance(raw_value, datetime):
        return raw_value
    if isinstance(raw_value, date):
        return datetime(raw_value.year, raw_value.month, raw_value.day)

    # Número serial de Excel (rango típico: 1900-2100)
    if isinstance(raw_value, (int, float)) and 10000 < raw_value < 80000:
        try:
            from openpyxl.utils.datetime import from_excel
            result = from_excel(raw_value)
            if isinstance(result, datetime):
                return result
            if isinstance(result, date):
                return datetime(result.year, result.month, result.day)
        except Exception:
            pass

    raw_str = str(raw_value).strip()

    # ── Formato MM/YYYY (sin día): "03/2026", "Período: 03/2026", etc. ──
    m = re.search(r"\b(0?[1-9]|1[0-2])/(20\d{2})\b", raw_str)
    if m:
        try:
            return datetime(int(m.group(2)), int(m.group(1)), 1)
        except ValueError:
            pass

    # ── Formato YYYY-MM (sin día): "2026-03" ─────────────────────────────
    m = re.search(r"\b(20\d{2})-(0[1-9]|1[0-2])\b", raw_str)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), 1)
        except ValueError:
            pass

    # ── Fechas completas con día ─────────────────────────────────────────
    candidates = re.findall(
        r"\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}|\d{4}[/\-]\d{1,2}[/\-]\d{1,2}",
        raw_str
    )
    if not candidates:
        candidates = [raw_str]
    else:
        candidates = candidates + [raw_str]

    for candidate in candidates:
        for fmt in (
            "%d/%m/%Y",   # 10/04/2026 (latino, prioridad)
            "%m/%d/%Y",   # 4/10/2026  (Excel inglés)
            "%Y-%m-%d",   # 2026-04-10
            "%d-%m-%Y",   # 10-04-2026
            "%Y/%m/%d",   # 2026/04/10
            "%d/%m/%y",   # 10/04/26
            "%m/%d/%y",   # 4/10/26
        ):
            try:
                return datetime.strptime(candidate.strip(), fmt)
            except ValueError:
                continue

    return None


def read_date_from_cell(path: Path) -> datetime:
    """
    Lee la fecha de realización de un archivo individual.

    Estrategia en cascada:
      1. Celda configurada en DATE_CELL (config.py, default "H2").
      2. Escaneo de TODAS las celdas de las primeras 5 filas buscando
         una celda que sea fecha Y esté cerca de un texto que contenga
         "fecha" (como "Fecha de diligenciamiento:").
      3. Escaneo de todas las celdas de las primeras 5 filas con cualquier
         valor parseable como fecha.
      4. Extracción desde el nombre del archivo (patrón YYYYMM).
    """
    from config import INDIVIDUAL_REQUIRED_COLS

    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    all_sheets = wb.sheetnames

    # ── Diagnóstico: volcar primeras 5 filas de TODAS las hojas ─────────
    logger.info(f"  [DIAG] Hojas en el archivo: {all_sheets}")
    cell_map = {}  # (sheet, row, col) → value
    for sname in all_sheets:
        ws_tmp = wb[sname]
        for row in ws_tmp.iter_rows(min_row=1, max_row=5, values_only=False):
            for cell in row:
                if cell.value is not None:
                    cell_map[(sname, cell.row, cell.column)] = cell.value
                    logger.info(
                        f"    [{sname}] {cell.coordinate}: {repr(cell.value)} "
                        f"({type(cell.value).__name__})"
                    )

    # ── Leer DATE_CELL desde la hoja de DATOS (no la portada) ───────────
    # Detectar hoja de datos igual que read_individual_data
    required_norm = [normalize_value(c) for c in INDIVIDUAL_REQUIRED_COLS]
    best_sheet = all_sheets[0]
    best_score = 0
    for sname in all_sheets:
        ws_tmp = wb[sname]
        _, score = _score_sheet_header(ws_tmp, required_norm)
        if score > best_score:
            best_score = score
            best_sheet = sname

    ws = wb[best_sheet]
    logger.info(f"  [FECHA] Leyendo {DATE_CELL} desde hoja de datos: '{best_sheet}'")

    # ── Estrategia 1: celda configurada ─────────────────────────────────
    target_cell = ws[DATE_CELL]
    raw = target_cell.value
    result = _parse_date_value(raw)
    if result:
        wb.close()
        logger.info(f"  [FECHA] Leída desde {DATE_CELL}: {result.strftime('%Y-%m-%d')}")
        return result

    logger.warning(f"  [FECHA] {DATE_CELL} vacía o inválida ({repr(raw)}). Buscando en hoja...")

    # ── Estrategia 2: buscar junto a etiqueta "fecha" o "period" ────────
    for (sname, r, c), val in cell_map.items():
        val_norm = normalize_value(val)
        if "fecha" in val_norm or "period" in val_norm:
            for dc in range(1, 6):
                neighbor = cell_map.get((sname, r, c + dc))
                parsed = _parse_date_value(neighbor)
                if parsed and parsed.year >= 2000:
                    wb.close()
                    ref = f"{get_column_letter(c + dc)}{r}"
                    logger.warning(
                        f"  [FECHA] Encontrada junto a '{val}' en [{sname}]{ref}: "
                        f"{parsed.strftime('%Y-%m-%d')}"
                    )
                    return parsed
            # También intentar parsear el propio valor de la celda (ej: "Período: 03/2026")
            parsed = _parse_date_value(val)
            if parsed and parsed.year >= 2000:
                wb.close()
                ref = f"{get_column_letter(c)}{r}"
                logger.warning(
                    f"  [FECHA] Extraída del texto '{val}' en [{sname}]{ref}: "
                    f"{parsed.strftime('%Y-%m-%d')}"
                )
                return parsed

    # ── Estrategia 3: cualquier celda parseable como fecha ───────────────
    for (sname, r, c), val in sorted(cell_map.items()):
        parsed = _parse_date_value(val)
        if parsed and parsed.year >= 2000:
            wb.close()
            ref = f"{get_column_letter(c)}{r}"
            logger.warning(
                f"  [FECHA] Primera fecha válida en [{sname}]{ref}: "
                f"{parsed.strftime('%Y-%m-%d')}"
            )
            return parsed

    wb.close()

    # ── Estrategia 4: extraer YYYYMM del nombre del archivo ─────────────
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", path.name)
    if match:
        year, month = int(match.group(1)), int(match.group(2))
        result = datetime(year, month, 1)
        logger.warning(
            f"  [FECHA] Extraída del nombre '{path.name}': "
            f"{result.strftime('%Y-%m-%d')} (día 1 del periodo)"
        )
        return result

    raise ValueError(
        f"No se encontró fecha válida en '{path.name}'. "
        f"Configura DATE_CELL en config.py con la celda correcta."
    )


# Alias para compatibilidad con main.py
read_date_from_h2 = read_date_from_cell


def read_execution_period(path: Path) -> int | None:
    """
    Lee el período de EJECUCIÓN (YYYYMM) directamente de la hoja de datos.

    Busca en las primeras 3 filas de la hoja de datos cualquier celda cuyo
    valor sea exactamente un número de 6 dígitos en formato YYYYMM
    (ej: 202603, '202603').

    Esto es distinto a la fecha de H2, que puede ser la fecha de
    diligenciamiento (cuando se llenó el formulario), no el período reportado.

    Retorna int (ej: 202603) o None si no encuentra nada.
    """
    from config import INDIVIDUAL_REQUIRED_COLS
    import re as _re

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_sheets = wb.sheetnames
    required_norm = [normalize_value(c) for c in INDIVIDUAL_REQUIRED_COLS]

    # Encontrar la hoja de datos
    best_sheet = all_sheets[0]
    best_score = 0
    for sname in all_sheets:
        ws_tmp = wb[sname]
        _, score = _score_sheet_header(ws_tmp, required_norm)
        if score > best_score:
            best_score = score
            best_sheet = sname

    ws = wb[best_sheet]
    periodo = None

    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for cell_val in row:
            if cell_val is None:
                continue
            raw = str(cell_val).strip()
            # Coincidencia exacta: 6 dígitos YYYYMM (ej: "202603")
            if _re.fullmatch(r"20\d{2}(0[1-9]|1[0-2])", raw):
                try:
                    periodo = int(raw)
                    break
                except ValueError:
                    pass
        if periodo:
            break

    wb.close()
    return periodo


# ─────────────────────────────────────────────────────────────
# CAMPOS DERIVADOS DE LA FECHA
# ─────────────────────────────────────────────────────────────

def build_date_fields(fecha: datetime) -> dict:
    anio = fecha.year
    mes_num = fecha.month
    periodo = int(f"{anio}{mes_num:02d}")
    mes = MESES_ES[mes_num] if MES_FORMAT == "texto" else mes_num
    return {"Fecha": fecha, "Anio": anio, "Mes": mes, "Periodo_YYYYMM": periodo}


# ─────────────────────────────────────────────────────────────
# ARCHIVO INDIVIDUAL
# ─────────────────────────────────────────────────────────────

def _score_sheet_header(ws, required_norm: list[str], max_scan: int = 20) -> tuple[int, int]:
    """
    Evalúa una hoja: retorna (mejor_fila_idx, mejor_score).
    Score = número de columnas requeridas encontradas en esa fila.
    """
    best_row = 0
    best_score = 0

    for row_idx, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True)):
        score = 0
        for cell_val in row:
            if cell_val is None:
                continue
            cell_norm = normalize_value(cell_val)
            for req_norm in required_norm:
                if cell_norm == req_norm or cell_norm.startswith(req_norm):
                    score += 1
                    break
        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row, best_score


def _detect_sheet_and_header(
    path: Path, required_cols: list[str], max_scan: int = 20
) -> tuple[str, int]:
    """
    Escanea TODAS las hojas del archivo para encontrar la que contiene
    la tabla de datos, y dentro de ella la fila de encabezado.

    Retorna (nombre_hoja, fila_header_base0).
    """
    required_norm = [normalize_value(c) for c in required_cols]

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    best_sheet = sheet_names[0]
    best_row = INDIVIDUAL_HEADER_ROW
    best_score = 0

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        row_idx, score = _score_sheet_header(ws, required_norm, max_scan)
        logger.info(
            f"  [HOJA] '{sheet_name}' → score={score} "
            f"(fila candidata: {row_idx + 1})"
        )
        if score > best_score:
            best_score = score
            best_sheet = sheet_name
            best_row = row_idx

    wb.close()

    if best_score == 0:
        logger.warning(
            f"  [HOJA] No se detectó hoja con tabla de datos. "
            f"Usando primera hoja, fila {INDIVIDUAL_HEADER_ROW + 1}."
        )
        return sheet_names[0], INDIVIDUAL_HEADER_ROW

    logger.info(
        f"  [HOJA] Seleccionada: '{best_sheet}' | "
        f"Encabezado en fila {best_row + 1} | Score={best_score}"
    )
    return best_sheet, best_row


def _normalize_column_names(df: pd.DataFrame, required_cols: list[str]) -> pd.DataFrame:
    """
    Renombra columnas del DataFrame para que coincidan con los nombres requeridos.

    Regla: si normalize(columna_actual) comienza con normalize(columna_requerida),
    se renombra a la columna requerida.

    Ejemplo:  "Ejecución *"  →  "Ejecucion"
              "Equipo "      →  "Equipo"
    """
    rename_map = {}
    for col in df.columns:
        col_norm = normalize_value(str(col))
        for req in required_cols:
            req_norm = normalize_value(req)
            if col_norm == req_norm or col_norm.startswith(req_norm):
                if col != req:
                    rename_map[col] = req
                    logger.info(f"  [COLUMNA] Renombrada: '{col}' → '{req}'")
                break

    if rename_map:
        df = df.rename(columns=rename_map)
    return df


def read_individual_data(path: Path) -> pd.DataFrame:
    """
    Lee la tabla de datos de un archivo individual con:
    - Auto-detección de la hoja correcta (escanea todas las hojas)
    - Auto-detección de la fila de encabezado
    - Normalización de nombres de columnas
    - Limpieza de filas y columnas vacías
    """
    from config import INDIVIDUAL_REQUIRED_COLS

    logger.debug(f"Leyendo datos de individual: {path.name}")

    # Si INDIVIDUAL_SHEET está configurado manualmente, usarlo sin escanear
    if INDIVIDUAL_SHEET is not None:
        sheet = INDIVIDUAL_SHEET
        from excel_reader import _score_sheet_header
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if isinstance(sheet, str) else wb[wb.sheetnames[sheet]]
        header_row, _ = _score_sheet_header(ws, [normalize_value(c) for c in INDIVIDUAL_REQUIRED_COLS])
        wb.close()
    else:
        sheet, header_row = _detect_sheet_and_header(path, INDIVIDUAL_REQUIRED_COLS)

    df = pd.read_excel(
        path,
        sheet_name=sheet,
        header=header_row,
        engine="openpyxl"
    )

    # Eliminar columnas completamente vacías o "Unnamed"
    df = df.loc[:, ~df.columns.astype(str).str.match(r"^Unnamed")]
    df = df.dropna(axis=1, how="all")

    # Eliminar filas completamente vacías
    df = df.dropna(how="all").reset_index(drop=True)

    # Normalizar nombres de columnas (ej: "Ejecución *" → "Ejecucion")
    df = _normalize_column_names(df, INDIVIDUAL_REQUIRED_COLS)

    # Eliminar filas donde faltan columnas clave (Equipo, Indicador o Ejecucion).
    # Esto filtra filas de notas/instrucciones que tienen texto solo en una columna.
    key_cols_present = [c for c in ("Equipo", "Indicador", "Ejecucion") if c in df.columns]
    if key_cols_present:
        mask_validas = df[key_cols_present].notna().all(axis=1)
        n_antes = len(df)
        df = df[mask_validas].reset_index(drop=True)
        n_filtradas = n_antes - len(df)
        if n_filtradas > 0:
            logger.info(
                f"  → {n_filtradas} fila(s) descartada(s) por tener "
                f"columnas clave vacías (filas de notas o instrucciones)."
            )

    if df.empty:
        raise ValueError(f"El archivo individual está vacío: {path.name}")

    logger.info(f"  Columnas encontradas: {list(df.columns)}")
    logger.info(f"  → {len(df)} registros leídos.")
    return df


def validate_individual_columns(df: pd.DataFrame, required: list[str], filename: str) -> list[str]:
    """
    Retorna columnas requeridas que faltan en el DataFrame.
    """
    missing = [col for col in required if col not in df.columns]
    if missing:
        logger.warning(
            f"  Columnas requeridas no encontradas en '{filename}': {missing}. "
            f"Columnas disponibles: {list(df.columns)}"
        )
    return missing


def validate_maestro_columns(df: pd.DataFrame, required: list[str]) -> list[str]:
    return [col for col in required if col not in df.columns]
