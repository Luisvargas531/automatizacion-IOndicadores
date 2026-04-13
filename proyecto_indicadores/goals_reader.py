"""
goals_reader.py
─────────────────────────────────────────────────────────────────────────────
Lee archivos de metas anuales por equipo.

Responsabilidades:
  - Detectar la hoja de metas (por defecto "🎯 Metas Anuales", con fallback
    por substring normalizado o cualquier hoja que contenga "meta")
  - Parsear columnas "Meta Enero *" → "Meta Diciembre *" y canonizarlas
    como Meta_01 … Meta_12 para uso interno
  - Extraer el año del nombre del archivo (patrón YYYYMM o solo YYYY)
  - Retornar un DataFrame normalizado listo para goals_updater
─────────────────────────────────────────────────────────────────────────────
"""

import logging
import re
from pathlib import Path

import openpyxl
import pandas as pd

from normalization import normalize_value

logger = logging.getLogger("indicadores")

# Mapeo número de mes → variantes de nombre reconocibles en columnas
_MES_NOMBRES: dict[int, list[str]] = {
    1:  ["enero", "jan", "january"],
    2:  ["febrero", "feb", "february"],
    3:  ["marzo", "mar", "march"],
    4:  ["abril", "abr", "april"],
    5:  ["mayo", "may"],
    6:  ["junio", "jun", "june"],
    7:  ["julio", "jul", "july"],
    8:  ["agosto", "ago", "august"],
    9:  ["septiembre", "sep", "sept", "september"],
    10: ["octubre", "oct", "october"],
    11: ["noviembre", "nov", "november"],
    12: ["diciembre", "dic", "dec", "december"],
}

# Nombre canónico de columna por mes (usado en el DataFrame de salida)
MES_CANONICAL: dict[int, str] = {n: f"Meta_{n:02d}" for n in range(1, 13)}


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS INTERNOS
# ─────────────────────────────────────────────────────────────────────────────

def extract_year_from_filename(filename: str) -> int | None:
    """Extrae el año (4 dígitos, 20xx o 19xx) del nombre del archivo."""
    match = re.search(r"(19|20)\d{2}", filename)
    return int(match.group(0)) if match else None


def extract_equipo_from_filename(filename: str) -> str | None:
    """
    Extrae el nombre del equipo del nombre del archivo de metas.

    Patrones esperados:
      METAS_ANALISIS_CUANTITATIVO_2026.xlsx → "ANALISIS CUANTITATIVO"
      METAS_CANALES_Y_SERVICIOS_2026.xlsx   → "CANALES Y SERVICIOS"

    Estrategia:
      1. Elimina prefijo "METAS_" (insensible a mayúsculas)
      2. Elimina sufijo "_YYYY" y la extensión
      3. Reemplaza "_" por espacio
    """
    stem = Path(filename).stem            # sin extensión
    stem_upper = stem.upper()

    # Quitar prefijo METAS_ si existe
    if stem_upper.startswith("METAS_"):
        stem = stem[6:]                   # longitud de "METAS_"

    # Quitar sufijo _YYYY (año de 4 dígitos al final)
    stem = re.sub(r"_(19|20)\d{2}$", "", stem)

    equipo = stem.replace("_", " ").strip()
    return equipo if equipo else None


def _detect_meta_sheet(path: Path, sheet_hint: str) -> str | None:
    """
    Busca la hoja de metas en el archivo:
      1. Coincidencia exacta con sheet_hint
      2. Substring normalizado de sheet_hint
      3. Cualquier hoja que contenga "meta" en el nombre
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    if sheet_hint in sheet_names:
        return sheet_hint

    hint_norm = normalize_value(sheet_hint)
    for name in sheet_names:
        if hint_norm in normalize_value(name):
            return name

    for name in sheet_names:
        if "meta" in normalize_value(name):
            return name

    logger.warning(
        f"  [METAS] No se encontró hoja '{sheet_hint}' en '{path.name}'. "
        f"Hojas disponibles: {sheet_names}"
    )
    return None


def _map_meta_columns(df: pd.DataFrame) -> dict[int, str]:
    """
    Identifica columnas del DataFrame que corresponden a cada mes.
    Busca columnas cuyo nombre normalizado contenga "meta" + algún nombre de mes.

    Retorna dict: {mes_número: nombre_columna_original}
    """
    col_map: dict[int, str] = {}
    for col in df.columns:
        col_norm = normalize_value(str(col))
        if "meta" not in col_norm:
            continue
        for mes_num, nombres in _MES_NOMBRES.items():
            if mes_num in col_map:
                continue
            for nombre in nombres:
                if nombre in col_norm:
                    col_map[mes_num] = col
                    break
    return col_map


def _detect_header_row(path: Path, sheet_name: str, max_scan: int = 20) -> int:
    """
    Escanea las primeras `max_scan` filas de la hoja y retorna el índice (0-based)
    de la fila que más se parece a un encabezado de metas.

    Busca palabras clave: "indicador", "equipo", "meta", "periodicidad".
    """
    keywords = {"indicador", "equipo", "meta", "periodicidad", "anual"}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    best_row = 0
    best_score = -1

    for row_idx, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True)):
        score = 0
        for cell in row:
            if cell is None:
                continue
            cell_norm = normalize_value(str(cell))
            for kw in keywords:
                if kw in cell_norm:
                    score += 1
                    break
        if score > best_score:
            best_score = score
            best_row = row_idx

    wb.close()
    return best_row


def _find_col(df: pd.DataFrame, keywords: list[str], exact: bool = False) -> str | None:
    """Busca la primera columna cuyo nombre normalizado contenga todos los keywords."""
    for col in df.columns:
        col_norm = normalize_value(str(col))
        if exact:
            if col_norm == keywords[0]:
                return col
        else:
            if all(kw in col_norm for kw in keywords):
                return col
    return None


# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def read_meta_file(path: Path, sheet_hint: str) -> tuple[pd.DataFrame, int | None]:
    """
    Lee un archivo de metas anuales y retorna un DataFrame normalizado.

    Parámetros
    ----------
    path       : Path  Ruta al archivo de metas.
    sheet_hint : str   Nombre esperado de la hoja (ej. "🎯 Metas Anuales").

    Retorna
    -------
    tuple[pd.DataFrame, int | None]
        - DataFrame con columnas:
            Equipo, Indicador, Periodicidad, Meta_Anual, Meta_01 … Meta_12
          (Meta_0N es None si el mes no aparece en el archivo).
        - Año extraído del nombre del archivo, o None si no se pudo.

    Lanza
    -----
    ValueError  Si no se encuentra la hoja de metas o las columnas mínimas.
    """
    anio = extract_year_from_filename(path.name)
    if anio is None:
        logger.warning(
            f"  [METAS] No se pudo extraer el año de '{path.name}'. "
            f"Se procesará sin validación de año."
        )

    sheet_name = _detect_meta_sheet(path, sheet_hint)
    if sheet_name is None:
        raise ValueError(
            f"No se encontró ninguna hoja de metas en '{path.name}'."
        )

    # Auto-detectar fila de encabezado (puede que el archivo tenga título en fila 1)
    header_row = _detect_header_row(path, sheet_name)
    logger.info(
        f"  [METAS] Leyendo hoja '{sheet_name}' de '{path.name}' "
        f"(encabezado en fila {header_row + 1})"
    )

    raw_df = pd.read_excel(path, sheet_name=sheet_name, header=header_row)
    raw_df.dropna(how="all", axis=1, inplace=True)
    raw_df.dropna(how="all", axis=0, inplace=True)
    raw_df.reset_index(drop=True, inplace=True)

    # Detectar columnas por rol
    equipo_col     = _find_col(raw_df, ["equipo"], exact=True)
    indicador_col  = _find_col(raw_df, ["indicador"])
    periodicidad_col = _find_col(raw_df, ["periodicidad"])
    meta_anual_col = _find_col(raw_df, ["meta", "anual"]) or \
                     _find_col(raw_df, ["meta", "annual"]) or \
                     _find_col(raw_df, ["meta", "total"])

    # Si no hay columna Equipo, intentar extraerlo del nombre del archivo
    equipo_from_filename: str | None = None
    if equipo_col is None:
        equipo_from_filename = extract_equipo_from_filename(path.name)
        if equipo_from_filename:
            logger.info(
                f"  [METAS] Columna 'Equipo' no encontrada. "
                f"Se usará el nombre extraído del archivo: '{equipo_from_filename}'"
            )
        else:
            raise ValueError(
                f"'{path.name}' no tiene columna 'Equipo' y no se pudo "
                f"extraer el equipo del nombre del archivo. "
                f"Columnas encontradas: {list(raw_df.columns)}"
            )

    if indicador_col is None:
        raise ValueError(
            f"'{path.name}' no tiene columna 'Indicador'. "
            f"Columnas encontradas: {list(raw_df.columns)}"
        )

    col_map_mes = _map_meta_columns(raw_df)
    logger.info(
        f"  [METAS] Columnas de mes detectadas: "
        f"{ {k: v for k, v in sorted(col_map_mes.items())} }"
    )

    # Construir DataFrame normalizado
    result_rows = []
    for _, row in raw_df.iterrows():
        equipo    = equipo_from_filename if equipo_col is None else row.get(equipo_col, "")
        indicador = row.get(indicador_col, "")

        # Omitir filas vacías o de encabezado repetido
        if pd.isna(equipo) or pd.isna(indicador):
            continue
        equipo    = str(equipo).strip()
        indicador = str(indicador).strip()
        if not equipo or not indicador:
            continue

        # Omitir filas de notas/instrucciones: el indicador es demasiado largo
        # para ser un nombre real (umbral: 100 caracteres)
        if len(indicador) > 100:
            logger.debug(
                f"  [METAS] Fila descartada (indicador muy largo, posible nota): "
                f"'{indicador[:60]}...'"
            )
            continue

        periodicidad = ""
        if periodicidad_col:
            v = row.get(periodicidad_col, "")
            periodicidad = "" if pd.isna(v) else str(v).strip()

        meta_anual = None
        if meta_anual_col:
            v = row.get(meta_anual_col, None)
            meta_anual = None if (v is None or (isinstance(v, float) and pd.isna(v))) else v

        record: dict = {
            "Equipo":        equipo,
            "Indicador":     indicador,
            "Periodicidad":  periodicidad,
            "Meta_Anual":    meta_anual,
        }

        for mes_num in range(1, 13):
            canonical = MES_CANONICAL[mes_num]
            if mes_num in col_map_mes:
                v = row.get(col_map_mes[mes_num], None)
                record[canonical] = None if (v is None or (isinstance(v, float) and pd.isna(v))) else v
            else:
                record[canonical] = None

        result_rows.append(record)

    goals_df = pd.DataFrame(result_rows)
    logger.info(
        f"  [METAS] {len(goals_df)} indicador(es) leído(s) de '{path.name}' (año={anio})"
    )
    return goals_df, anio
